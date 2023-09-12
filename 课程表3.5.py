# import
from PIL import Image, ImageDraw, ImageFont # create image
from datetime import datetime # get time
from time import sleep, time # sleep
from ctypes import windll # set wallpaper
from openpyxl import load_workbook # open excel spreadsheet
from os import getcwd # get current path

# constants in CAPITAL LETTERS :)
CURRENT_PATH = getcwd()
SETTINGS = {"刷新频率": 1,
            "图片清晰度": 0,
            "文本描边粗细": 0,
            "背景颜色": "FFF5E1",
            "文本颜色": "202020",
            "文本描边颜色": "FFFFFF"}

# functions
def set_wallpaper(image_path):
    windll.user32.SystemParametersInfoW(20, 0, image_path, 0)
def hex_to_rgb(hex_value):
    return tuple(int(hex_value[i:i+2], 16) for i in (0, 2, 4))
def number_to_time(number):
    minute = number // 60
    hour = minute // 60
    minute = minute % 60
    return f"{str(hour).rjust(2,'0')}:{str(minute).rjust(2,'0')}"
def draw_text(y,text,size):
    global image,draw,SETTINGS
    font = ImageFont.truetype(SETTINGS["FONT"], size)
    _, _, w, h = draw.textbbox((0, 0), text, font=font)
    draw.text(((SETTINGS["WIDTH"]-w)/2, y), text, font=font, fill=SETTINGS["文本颜色"], stroke_width=SETTINGS["文本描边粗细"], stroke_fill=SETTINGS["文本描边颜色"])
def draw_text_ext(x,y,text,size,color,align=0):
    global image,draw,SETTINGS
    font = ImageFont.truetype(SETTINGS["FONT"], size)
    _, _, w, h = draw.textbbox((0, 0), text, font=font)
    if align == 0: draw.text((x, y), text, font=font, fill=color)
    elif align == 1: draw.text((x-w/2, y), text, font=font, fill=color)
    elif align == 2: draw.text((x-w, y), text, font=font, fill=color)
#def debug():
#    global now
#    pass
#    now = datetime(now.year, now.month, now.day, now.hour-1, now.minute, now.second, now.microsecond)

def reload():
    global today_variant,wholeday_schedule,bg,today_text,pointer,pointer_next,schedule,SETTINGS
    
    # read excel
    excel = load_workbook("settings.xlsx",data_only=True)
    #  read semester schedule
    semester_schedule_sheet = excel["课程安排"]
    now = datetime.today()
    ind = 2
    while True:
        key = semester_schedule_sheet.cell(ind,1).value
        if key == None: break
        if 0 <= (now-key).days < 7:
            if semester_schedule_sheet.cell(ind,(now-key).days+2).value == None:
                today_variant = semester_schedule_sheet.cell(1,(now-key).days+2).value # variant means the day's schedule
            else: today_variant = semester_schedule_sheet.cell(ind,(now-key).days+2).value
            break
        ind += 1
    #  read text display
    semester_schedule_sheet = excel["文本显示"]
    ind = 2
    while True:
        key = semester_schedule_sheet.cell(ind,1).value
        if key == None: break
        if 0 <= (now-key).days < 7:
            if semester_schedule_sheet.cell(ind,(now-key).days+2).value == None:
                today_text = ""
            else: today_text = semester_schedule_sheet.cell(ind,(now-key).days+2).value
            break
        ind += 1
    #  read settings
    settings_sheet = excel["设置"]
    ind = 1 # 1-indexing
    while True: 
        key = settings_sheet.cell(ind, 1).value
        if key == None: break
        else: value = settings_sheet.cell(ind, 2).value
        if key in {"刷新频率","图片清晰度","文本描边粗细"}:
            try: SETTINGS[key] = value
            except: pass
        elif key in {"背景颜色","文本颜色","文本描边颜色"}:
            try: SETTINGS[key] = hex_to_rgb(value)
            except: pass
        else: SETTINGS[key] = value
        ind += 1
    #  read advanced settings
    advanced_settings_sheet = excel["高级设置"]
    ind = 1 # 1-indexing
    while True: 
        key = advanced_settings_sheet.cell(ind, 1).value
        if key == None: break
        else: value = advanced_settings_sheet.cell(ind, 2).value
        if value == None: value = ""
        SETTINGS[key] = value
        ind += 1
    #  modify advanced settings
    SETTINGS["AFTERSCHOOL"] = SETTINGS["AFTERSCHOOL"].split()
    SETTINGS["BREAK"] = SETTINGS["BREAK"].split()
    SETTINGS["CLASS"] = SETTINGS["CLASS"].split()
    SETTINGS["EMPTYCLASS"] = SETTINGS["EMPTYCLASS"].split()
    SETTINGS["COLORLIGHT"] = hex_to_rgb(SETTINGS["COLORLIGHT"])
    #  read schedule
    schedule_sheet = excel["课表"]
    day_variant_count = 0
    while schedule_sheet.cell(1, day_variant_count+2).value != None:
        day_variant_count += 1 # get the day variants
    schedule = {}
    wholeday_schedule = {}
    for d in range(day_variant_count): # day
        ind = 2 # the 1st line is 星期一blablabla
        day_variant = schedule_sheet.cell(1, d+2).value
        schedule[day_variant] = []
        wholeday_schedule[day_variant] = []
        while True:
            value = schedule_sheet.cell(ind, d+2).value
            if value == None:
                ind += 1
                continue
            t = schedule_sheet.cell(ind, 1).value # time
            h = t.hour # hour
            m = t.minute # minute
            s = t.second # second
            schedule[day_variant].append((h*60*60+m*60+s, value))
            ind += 1
            if value in SETTINGS["CLASS"] or value in SETTINGS["EMPTYCLASS"]:
                wholeday_schedule[day_variant].append(value)
            if value in SETTINGS["AFTERSCHOOL"]: # break only if afterschool
                break

    # update pointer
    #now = datetime.today()
    #debug()
    now_value = (now.hour*60+now.minute)*60+now.second
    pointer = 0
    while True:
        if pointer+1 == len(schedule[today_variant]):
            break
        if schedule[today_variant][pointer+1][0] <= now_value:
            pointer += 1
        else:
            break
    pointer_next = pointer + 1 # another pointer showing the next class
    while pointer_next <= pointer or (schedule[today_variant][pointer][1] in SETTINGS["BREAK"] and schedule[today_variant][pointer_next][1] in SETTINGS["BREAK"]):
        # we dont need to prevent overflow here, it's super safe :)
        pointer_next += 1

    # generate bg
    image1 = Image.new("RGB",(SETTINGS["WIDTH"],SETTINGS["HEIGHT"]),SETTINGS["背景颜色"])
    try:
        bg = Image.open("background.png")
        bg = bg.resize((SETTINGS["WIDTH"],SETTINGS["HEIGHT"]))
        bg = Image.blend(image1,bg,SETTINGS["图片清晰度"])
    except: bg = image1

reload()
prev = datetime.today()
# main loop
while True:
    now = datetime.today()
    if (now.year,now.month,now.day,now.minute) != (prev.year,prev.month,prev.day,prev.minute):
        prev = datetime.today()
        reload()
    #debug()
    now_value = (now.hour*60+now.minute)*60+now.second
    image = bg.copy()
    draw = ImageDraw.Draw(image)
    # draw the whole schedule
    for ind,item in enumerate(wholeday_schedule[today_variant]):
        if item in SETTINGS["CLASS"]:
            draw_text_ext(SETTINGS["POSSCHEDULEX"],SETTINGS["POSSCHEDULEY"]+ind*(SETTINGS["SIZESMALL"]+SETTINGS["SPACESCHEDULE"]),item,SETTINGS["SIZESMALL"],SETTINGS["文本颜色"])
        elif item in SETTINGS["EMPTYCLASS"]:
            draw_text_ext(SETTINGS["POSSCHEDULEX"],SETTINGS["POSSCHEDULEY"]+ind*(SETTINGS["SIZESMALL"]+SETTINGS["SPACESCHEDULE"]),item,SETTINGS["SIZESMALL"],SETTINGS["COLORLIGHT"])
    # draw some text at the upper right corner
    ind = 0
    draw_text_ext(SETTINGS["WIDTH"]-SETTINGS["POSSCHEDULEX"],SETTINGS["POSSCHEDULEY"]+ind*(SETTINGS["SIZESMALL"]+SETTINGS["SPACESCHEDULE"]),SETTINGS["TEXT"],SETTINGS["SIZESMALL"],SETTINGS["文本颜色"],2)
    ind += 1
    draw_text_ext(SETTINGS["WIDTH"]-SETTINGS["POSSCHEDULEX"],SETTINGS["POSSCHEDULEY"]+ind*(SETTINGS["SIZESMALL"]+SETTINGS["SPACESCHEDULE"]),today_text,SETTINGS["SIZESMALL"],SETTINGS["文本颜色"],2)
    
    # draw the time now
    draw_text(SETTINGS["POSSMALL"],f"{now.year}/{now.month}/{now.day} {str(now.hour).rjust(2,'0')}:{str(now.minute).rjust(2,'0')}:{str(now.second).rjust(2,'0')}",SETTINGS["SIZESMALL"])
            
    if pointer+1 != len(schedule[today_variant]): # testing for next event
        if schedule[today_variant][pointer+1][0] <= now_value:
            pointer += 1
            while pointer_next <= pointer or (schedule[today_variant][pointer][1] in SETTINGS["BREAK"] and schedule[today_variant][pointer_next][1] in SETTINGS["BREAK"]):
                pointer_next += 1 # safe to do this
    # draw the title
    draw_text(SETTINGS["POSTITLE"],schedule[today_variant][pointer][1],SETTINGS["SIZETITLE"])
    # draw the subtitle
    if schedule[today_variant][pointer][1] in SETTINGS["AFTERSCHOOL"]:
        subtitle = SETTINGS["CLASSOVER"]
    else:
        subtitle = ""
        delta_time = schedule[today_variant][pointer_next][0]-now_value
        delta_minute = delta_time // 60
        delta_second = delta_time % 60
        if delta_minute == 0:
            subtitle += f"{delta_second}秒"
        elif delta_second == 0:
            subtitle += f"{delta_minute}分钟"
        else:
            subtitle += f"{delta_minute}分{delta_second}秒"
            
        if schedule[today_variant][pointer_next][1] in SETTINGS["CLASS"]: # next = class
            subtitle += f"后上{schedule[today_variant][pointer_next][1]}课"
        elif schedule[today_variant][pointer][1] in SETTINGS["CLASS"] and schedule[today_variant][pointer_next][1] in SETTINGS["BREAK"]: # now = class and next = break
            if not SETTINGS["CLASSTIMER"]:
                subtitle = number_to_time(schedule[today_variant][pointer][0])+"~"+number_to_time(schedule[today_variant][pointer_next][0])
            else:
                subtitle += "后下课"
        else: # now = break and next = break
            subtitle += f"后{schedule[today_variant][pointer_next][1]}"
    draw_text(SETTINGS["POSSUBTITLE"],subtitle,SETTINGS["SIZESUBTITLE"])
    draw_text(SETTINGS["POSTINY"],SETTINGS["NOTE"],SETTINGS["SIZETINY"])

    try: image.save(f"{getcwd()}\\file\\bg.jpg")
    except: pass
    set_wallpaper(f"{getcwd()}\\file\\bg.jpg")

    sleep(SETTINGS["刷新频率"]-time()%SETTINGS["刷新频率"])
